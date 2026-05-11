import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from database import *
from datetime import datetime, date
import io

try:
    from fpdf import FPDF
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False

# ── Constantes ────────────────────────────────────────────────
MESES_PT  = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
             7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}
MESES_EN  = {1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
             7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"}
STATUS_PT = {"ativa":"Ativa","inativa":"Inativa","suspensa":"Suspensa"}
STATUS_EN = {"ativa":"Active","inativa":"Inactive","suspensa":"Suspended"}

PALETAS = {
    "normal":       {"completo":"#10b981","parcial":"#f59e0b","vazio":"#ef4444"},
    "deuteranopia": {"completo":"#0ea5e9","parcial":"#f59e0b","vazio":"#7c3aed"},
    "protanopia":   {"completo":"#0ea5e9","parcial":"#fbbf24","vazio":"#6366f1"},
    "tritanopia":   {"completo":"#10b981","parcial":"#f472b6","vazio":"#6366f1"},
}

# ── Page Config ───────────────────────────────────────────────
st.set_page_config(page_title="Gestor de Empresas", page_icon="🏢",
                   layout="wide", initial_sidebar_state="expanded")

# ── Settings init ─────────────────────────────────────────────
if "cfg" not in st.session_state:
    st.session_state.cfg = {
        "tema":      "claro",
        "idioma":    "pt",
        "fonte":     "medio",
        "daltonico": "normal",
    }
cfg = st.session_state.cfg

# ── Init DB ───────────────────────────────────────────────────
init_database()

# ── Dynamic CSS ───────────────────────────────────────────────
_FONTES   = {"pequeno":"13px","medio":"15px","grande":"18px"}
_fonte_px = _FONTES.get(cfg["fonte"], "15px")
_ESCURO   = cfg["tema"] == "escuro"

_bg_main    = "#0f172a" if _ESCURO else "#ffffff"
_bg_sidebar = "#1e293b" if _ESCURO else "#f8fafc"
_bg_card    = "#1e293b" if _ESCURO else "#f8fafc"
_border     = "#334155" if _ESCURO else "#e2e8f0"
_text_main  = "#f1f5f9" if _ESCURO else "#1e293b"
_text_sub   = "#94a3b8" if _ESCURO else "#64748b"
_text_tag   = "#cbd5e1" if _ESCURO else "#475569"
_bg_tag     = "#334155" if _ESCURO else "#e2e8f0"
_badge_row  = "#1e293b" if _ESCURO else "#f1f5f9"
_hover_row  = "#263148" if _ESCURO else "#f0f4ff"
_input_bg   = "#1e293b" if _ESCURO else "#ffffff"

_dark_overrides = f"""
    p, span, label, h1, h2, h3, h4, h5, h6 {{ color: {_text_main}; }}
    [data-testid="stMetricValue"], [data-testid="stMetricLabel"],
    [data-testid="stMetricDelta"] {{ color: {_text_main} !important; }}
    [data-testid="metric-container"] {{
        background-color: {_bg_card} !important;
        border-radius: 8px; padding: 6px;
    }}
    .stTabs [data-testid="stTabBar"] {{ background-color: {_bg_sidebar} !important; }}
    .stTabs [data-baseweb="tab"] {{ color: {_text_sub} !important; }}
    .stTabs [aria-selected="true"] {{ color: #3b82f6 !important; }}
    .stDataFrame, [data-testid="stDataFrame"] {{
        background-color: {_bg_card} !important;
    }}
    .stTextInput input, .stTextArea textarea, .stSelectbox div[data-baseweb="select"] {{
        background-color: {_input_bg} !important;
        color: {_text_main} !important;
        border-color: {_border} !important;
    }}
    .stExpander {{ background-color: {_bg_card} !important; border-color: {_border} !important; }}
    [data-testid="stForm"] {{ background-color: {_bg_card} !important; border-color: {_border} !important; }}
    .stAlert {{ background-color: {_bg_card} !important; }}
    [data-testid="stSidebar"] * {{ color: {_text_main}; }}
    .stRadio label, .stCheckbox label, .stSelectbox label {{ color: {_text_main} !important; }}
""" if _ESCURO else ""

st.markdown(f"""
<style>
/* ── font size global ── */
html, body, .stApp {{
    font-size: {_fonte_px} !important;
}}

/* ── background principal ── */
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMain"],
.main .block-container {{
    background-color: {_bg_main} !important;
}}
[data-testid="stSidebar"],
[data-testid="stSidebarContent"] {{
    background-color: {_bg_sidebar} !important;
}}

/* ── header ── */
#MainMenu, footer {{ visibility: hidden; }}
header[data-testid="stHeader"] {{ background: transparent !important; }}
header[data-testid="stHeader"] > div:first-child {{ visibility: hidden; }}
[data-testid="stExpandSidebarButton"],
[data-testid="stSidebarCollapseButton"],
[data-testid="collapsedControl"],
button[data-testid="baseButton-headerNoPadding"] {{ visibility: visible !important; }}

/* ── dark mode overrides ── */
{_dark_overrides}

/* ── cards de empresa ── */
.card {{
    background: {_bg_card};
    border: 1px solid {_border};
    border-radius: 12px;
    padding: 1.4rem 1.6rem;
    margin-bottom: 1rem;
    transition: border .2s, box-shadow .2s;
}}
.card:hover {{ border-color:#3b82f6; box-shadow:0 4px 16px rgba(59,130,246,.12); }}

.avatar {{
    width:48px; height:48px; border-radius:10px;
    background: linear-gradient(135deg,#60a5fa,#1e40af);
    display:inline-flex; align-items:center; justify-content:center;
    color:#fff; font-weight:700; font-size:18px; float:left; margin-right:12px;
}}
.company-name {{ font-size:16px; font-weight:700; color:{_text_main}; }}
.company-sub  {{ font-size:13px; color:{_text_sub}; }}

/* ── barra de progresso ── */
.prog-wrap {{ width:100%; background:{_border}; border-radius:4px; height:8px; margin:10px 0 4px; }}
.prog-bar  {{ height:8px; border-radius:4px; }}
.prog-pct  {{ font-size:18px; font-weight:700; }}

/* ── tags de etapa ── */
.tag {{
    display:inline-flex; align-items:center; gap:5px;
    background:{_bg_tag}; border-radius:6px; padding:5px 10px;
    font-size:12px; color:{_text_tag}; margin-right:6px; margin-top:4px;
}}

/* ── modal etapa row ── */
.etapa-row {{
    display:flex; align-items:center; gap:12px;
    background:{_badge_row}; border:1px solid {_border};
    border-radius:8px; padding:12px; margin-bottom:8px;
}}
.badge-ok {{ background:rgba(16,185,129,.2); color:#10b981;
             padding:3px 10px; border-radius:4px; font-size:12px; font-weight:600; }}
.badge-nd {{ background:{_bg_tag}; color:{_text_sub};
             padding:3px 10px; border-radius:4px; font-size:12px; }}

/* ── tabela de relatório ── */
.rel-table {{ width:100%; border-collapse:collapse; font-size:14px; }}
.rel-table th {{
    background:#3b82f6; color:#fff; padding:10px 14px;
    text-align:left; font-weight:600; font-size:12px;
    text-transform:uppercase; letter-spacing:.5px;
    position:sticky; top:0;
}}
.rel-table td {{
    padding:9px 14px; border-bottom:1px solid {_border};
    color:{_text_main}; background:{_bg_card};
}}
.rel-table tr:hover td {{ background:{_hover_row}; }}

/* ── mini progress bar no relatório ── */
.mini-prog-wrap {{ width:80px; background:{_border}; border-radius:3px; height:6px; display:inline-block; }}
.mini-prog-bar  {{ height:6px; border-radius:3px; display:block; }}

/* ── card de configurações ── */
.cfg-card {{
    background:{_bg_card}; border:1px solid {_border};
    border-radius:12px; padding:1.2rem 1.5rem; margin-bottom:1rem;
}}
.cfg-title {{
    font-size:15px; font-weight:700; color:{_text_main};
    margin-bottom:.8rem;
}}

/* ── Configurações fixo no rodapé da sidebar ── */
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {{
    display: flex !important;
    flex-direction: column !important;
    min-height: calc(100vh - 5rem) !important;
}}
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] > div:last-child {{
    margin-top: auto !important;
}}
/* estiliza o botão de configurações */
div[data-testid="stSidebar"] button[kind="secondary"]:last-of-type,
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] > div:last-child button {{
    background: transparent !important;
    border: 1px solid {_border} !important;
    color: {_text_sub} !important;
    font-size: 14px !important;
}}
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] > div:last-child button:hover {{
    border-color: #3b82f6 !important;
    color: #3b82f6 !important;
    background: rgba(59,130,246,.06) !important;
}}
</style>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────
st.title("🏢 Gestor de Empresas")

_is_pt = cfg["idioma"] == "pt"
_MESES = MESES_PT if _is_pt else MESES_EN
_STATUS = STATUS_PT if _is_pt else STATUS_EN

# ── Navigation state ─────────────────────────────────────────
# pagina_atual  → página de fato ativa (pode ser "⚙️ Configurações")
# _radio_page   → qual item do radio está selecionado (nunca "⚙️ Configurações")
_MENU_ITEMS = [
    "📋 Empresas",
    "➕ Cadastrar Empresa",
    "📥 Importar Planilha",
    "📈 Relatórios",
]
if "pagina_atual" not in st.session_state:
    st.session_state.pagina_atual = "📋 Empresas"
if "_radio_page" not in st.session_state:
    st.session_state._radio_page = "📋 Empresas"

with st.sidebar:
    st.markdown("---")
    _radio_sel = st.radio(
        "📋 Menu", _MENU_ITEMS,
        index=_MENU_ITEMS.index(st.session_state._radio_page),
    )
    # Só atualiza quando o usuário realmente clicar em outro item do radio
    if _radio_sel != st.session_state._radio_page:
        st.session_state._radio_page  = _radio_sel
        st.session_state.pagina_atual = _radio_sel
        st.rerun()

    st.markdown("---")

today = date.today()
mes_ref = st.sidebar.selectbox(
    "Mês de Referência" if _is_pt else "Reference Month",
    range(1, 13), index=today.month - 1,
    format_func=lambda x: _MESES[x]
)
ano_ref = st.sidebar.number_input(
    "Ano" if _is_pt else "Year",
    value=today.year, min_value=2020, max_value=2050
)
mes_str = f"{mes_ref:02d}"

# ── Botão Configurações — fixo no rodapé da sidebar ──────────
with st.sidebar:
    if st.button("⚙️ Configurações", key="btn_cfg_bottom", use_container_width=True):
        st.session_state.pagina_atual = "⚙️ Configurações"
        st.rerun()

menu = st.session_state.pagina_atual


# ── Helpers ───────────────────────────────────────────────────
def safe(v, fb=""):
    if v is None: return fb
    s = str(v).strip()
    return fb if s.lower() == "nan" else s

def initials(nome):
    words = nome.strip().split()
    return (words[0][0] + (words[1][0] if len(words) > 1 else "")).upper()

def progress_color(pct):
    paleta = PALETAS.get(cfg["daltonico"], PALETAS["normal"])
    if pct == 100: return paleta["completo"]
    if pct >= 50:  return paleta["parcial"]
    return paleta["vazio"]

def pct_emp(emp):
    total, done = emp[13], emp[14]
    return round(done / total * 100) if total > 0 else 0


# ══════════════════════════════════════════════════════════════
# DIALOG — Marcar etapa em lote
# ══════════════════════════════════════════════════════════════
@st.dialog("⚡ Ações em Lote", width="large")
def dialog_marcar_lote():
    todas_emps   = obter_todas_empresas()
    nomes_etapas = obter_nomes_etapas_distintos()

    opcoes_emps = {
        e[0]: f"{safe(e[10]) + ' · ' if e[10] else ''}{e[1]}"
        for e in todas_emps
    }

    def _selecionar_empresas(sufixo):
        """Widget reutilizável para escolher todas ou um subconjunto de empresas."""
        modo = st.radio(
            "Aplicar para",
            ["🏢 Todas as empresas", "☑️ Selecionar manualmente"],
            horizontal=True,
            key=f"modo_{sufixo}",
        )
        if modo == "☑️ Selecionar manualmente":
            ids = st.multiselect(
                "Empresas",
                options=list(opcoes_emps.keys()),
                format_func=lambda eid: opcoes_emps.get(eid, str(eid)),
                placeholder="Digite nome ou código para filtrar...",
                key=f"emps_{sufixo}",
            )
        else:
            ids = list(opcoes_emps.keys())
        return ids

    tab_marcar, tab_adicionar, tab_remover = st.tabs([
        "⚡ Marcar Status de Etapa",
        "➕ Adicionar Nova Etapa",
        "🗑️ Remover Etapa",
    ])

    # ══ TAB 1 — Marcar status ════════════════════════════════
    with tab_marcar:
        st.caption(f"Marca uma etapa já existente como concluída ou pendente — {_MESES[mes_ref]}/{ano_ref}")
        st.markdown("")

        if not nomes_etapas:
            st.warning("Nenhuma etapa cadastrada ainda.")
        else:
            col_e, col_s = st.columns([2, 1])
            with col_e:
                etapa_m = st.selectbox("📌 Etapa", nomes_etapas, key="m_etapa")
            with col_s:
                status_opt = st.radio(
                    "Status",
                    ["✅ Concluído", "❌ Pendente"],
                    horizontal=True,
                    key="m_status",
                )
            concluido = status_opt.startswith("✅")
            st.markdown("---")

            ids_m = _selecionar_empresas("marcar")
            n_m   = len(ids_m)

            if n_m > 0:
                acao = "concluída" if concluido else "pendente"
                st.info(
                    f"**{n_m}** empresa(s) terão **\"{etapa_m}\"** "
                    f"marcada como **{acao}** em {_MESES[mes_ref]}/{ano_ref}."
                )
            else:
                st.warning("Selecione pelo menos uma empresa.")

            st.markdown("")
            if st.button("⚡ Aplicar Status", type="primary",
                         use_container_width=True, disabled=(n_m == 0),
                         key="btn_marcar"):
                marcadas = marcar_etapa_bulk(ids_m, etapa_m, mes_str, ano_ref, concluido)
                st.success(
                    f"✅ **\"{etapa_m}\"** marcada como {acao} "
                    f"em **{marcadas}** empresa(s)!"
                )
                st.rerun()

    # ══ TAB 2 — Adicionar etapa ══════════════════════════════
    with tab_adicionar:
        st.caption("Adiciona uma etapa nova (ou existente) ao cadastro das empresas selecionadas.")
        st.markdown("")

        col_n, col_hint = st.columns([2, 1])
        with col_n:
            nova_etapa_nome = st.text_input(
                "Nome da etapa",
                placeholder="Ex: Folha de Pagamento, FGTS, IRPF...",
                key="add_etapa_nome",
            )
        with col_hint:
            if nomes_etapas:
                etapa_existente = st.selectbox(
                    "Ou use uma existente",
                    ["— digitar novo —"] + nomes_etapas,
                    key="add_etapa_existente",
                )
                if etapa_existente != "— digitar novo —":
                    nova_etapa_nome = etapa_existente

        st.markdown("---")
        ids_a = _selecionar_empresas("adicionar")
        n_a   = len(ids_a)

        nome_final = nova_etapa_nome.strip() if nova_etapa_nome else ""

        if nome_final and n_a > 0:
            st.info(
                f"A etapa **\"{nome_final}\"** será adicionada a "
                f"**{n_a}** empresa(s). Empresas que já possuem essa etapa serão ignoradas."
            )
        elif not nome_final:
            st.warning("Digite o nome da etapa.")
        else:
            st.warning("Selecione pelo menos uma empresa.")

        st.markdown("")
        if st.button("➕ Adicionar Etapa", type="primary",
                     use_container_width=True,
                     disabled=(not nome_final or n_a == 0),
                     key="btn_adicionar_etapa"):
            adicionadas, ja_existiam = adicionar_etapa_bulk(ids_a, nome_final)
            msg = f"✅ Etapa **\"{nome_final}\"** adicionada em **{adicionadas}** empresa(s)."
            if ja_existiam:
                msg += f" ({ja_existiam} já possuíam essa etapa e foram ignoradas.)"
            st.success(msg)
            st.rerun()

    # ══ TAB 3 — Remover etapa ════════════════════════════════
    with tab_remover:
        st.caption("Remove uma etapa do cadastro de empresas selecionadas.")
        st.markdown("")

        if not nomes_etapas:
            st.warning("Nenhuma etapa cadastrada para remover.")
        else:
            etapa_r = st.selectbox("📌 Selecione a etapa para remover", nomes_etapas, key="r_etapa")
            st.markdown("---")

            ids_r = _selecionar_empresas("remover")
            n_r   = len(ids_r)

            if n_r > 0:
                st.warning(
                    f"⚠️ A etapa **\"{etapa_r}\"** será **removida** de "
                    f"**{n_r}** empresa(s). Esta ação **não pode ser desfeita**!"
                )
            else:
                st.warning("Selecione pelo menos uma empresa.")

            st.markdown("")
            if st.button("🗑️ Remover Etapa", type="primary",
                         use_container_width=True, disabled=(n_r == 0),
                         key="btn_remover_etapa"):
                # Encontra o ID da etapa
                etapa_id = None
                todas_etapas = obter_nomes_etapas_distintos()
                # Busca a etapa por nome para pegar o ID
                for emp_id in ids_r:
                    etapas_emp = obter_etapas_empresa(emp_id)
                    for et in etapas_emp:
                        if et[2] == etapa_r:
                            etapa_id = et[0]
                            break
                    if etapa_id:
                        break

                if etapa_id:
                    removidas = 0
                    for emp_id in ids_r:
                        etapas_emp = obter_etapas_empresa(emp_id)
                        for et in etapas_emp:
                            if et[2] == etapa_r:
                                remover_etapa(et[0])
                                removidas += 1
                                break
                    st.success(
                        f"✅ Etapa **\"{etapa_r}\"** removida de **{removidas}** empresa(s)!"
                    )
                    st.rerun()


# ══════════════════════════════════════════════════════════════
# DIALOG — Detalhes da empresa
# ══════════════════════════════════════════════════════════════
@st.dialog("Detalhes da Empresa", width="large")
def dialog_empresa(empresa_id):
    emp       = obter_empresa_por_id(empresa_id)
    etapas    = obter_etapas_empresa(empresa_id)
    status_map = obter_status_etapas(empresa_id, mes_str, ano_ref)
    pct, done, total = calcular_progresso(empresa_id, mes_str, ano_ref)

    col_av, col_info = st.columns([1, 6])
    with col_av:
        st.markdown(
            f'<div class="avatar" style="font-size:22px;width:56px;height:56px;border-radius:12px;">'
            f'{initials(emp[1])}</div>', unsafe_allow_html=True)
    with col_info:
        st.markdown(f"**{emp[1]}**")
        sub = []
        if emp[10]: sub.append(f"Cód. {emp[10]}")
        if emp[5]:  sub.append(f"CNPJ: {emp[5]}")
        if emp[11]: sub.append(emp[11])
        st.caption(" · ".join(sub) if sub else "")

    cor = progress_color(pct)
    st.markdown(f"""
    <div style="display:flex;justify-content:space-between;align-items:center;margin:12px 0 4px;">
        <span style="font-size:12px;font-weight:600;color:#94a3b8;text-transform:uppercase;letter-spacing:.5px;">Progresso</span>
        <span style="font-size:20px;font-weight:700;color:{cor};">{pct}%</span>
    </div>
    <div class="prog-wrap"><div class="prog-bar" style="width:{pct}%;background:{cor};"></div></div>
    <div style="font-size:12px;color:#94a3b8;margin-bottom:16px;">{done}/{total} etapas concluídas</div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown("**ETAPAS**")
    for etapa in etapas:
        eid   = etapa[0]
        enome = etapa[2]
        feito = status_map.get(eid, False)

        col1, col2, col3 = st.columns([1, 6, 2])
        with col1:
            novo = st.checkbox("Concluído", value=feito, key=f"chk_{eid}_{mes_str}_{ano_ref}",
                               label_visibility="collapsed")
            if novo != feito:
                marcar_etapa(eid, mes_str, ano_ref, novo)
                st.rerun()
        with col2:
            label = f"~~{enome}~~" if feito else enome
            st.markdown(label)
        with col3:
            if feito:
                st.markdown('<span class="badge-ok">✓ Concluído</span>', unsafe_allow_html=True)
            else:
                st.markdown('<span class="badge-nd">⏳ Pendente</span>', unsafe_allow_html=True)

    st.divider()
    st.markdown("**Gerenciar Etapas**")

    # Layout: Adicionar e Remover lado a lado
    col_add, col_rem = st.columns(2)

    with col_add:
        st.markdown("##### ➕ Adicionar Nova Etapa")
        nova_etapa = st.text_input("Nova etapa", placeholder="Ex: Folha de Pagamento",
                                   label_visibility="collapsed", key=f"nova_etapa_{empresa_id}")
        if st.button("➕ Adicionar", key=f"btn_add_{empresa_id}", use_container_width=True):
            if nova_etapa.strip():
                adicionar_etapa(empresa_id, nova_etapa.strip())
                st.rerun()
            else:
                st.warning("Digite o nome da etapa.")

    with col_rem:
        st.markdown("##### 🗑️ Remover Etapa")
        etapas_atual = obter_etapas_empresa(empresa_id)
        nomes = [e[2] for e in etapas_atual]

        if not nomes:
            st.info("Nenhuma etapa para remover")
        else:
            remover = st.selectbox("Selecione a etapa", ["—"] + nomes,
                                   key=f"rem_{empresa_id}", label_visibility="collapsed")
            if st.button("🗑️ Remover", key=f"confirm_rem_{empresa_id}", type="primary",
                        use_container_width=True, disabled=(remover == "—")):
                for e in etapas_atual:
                    if e[2] == remover:
                        remover_etapa(e[0])
                        st.rerun()

    st.divider()
    st.markdown("**INFORMAÇÕES / OBSERVAÇÕES**")
    info_atual = emp[12] if len(emp) > 12 else ""
    info_texto = st.text_area("Informações", value=info_atual or "",
                              placeholder="Anote aqui informações relevantes...",
                              height=120, label_visibility="collapsed",
                              key=f"info_{empresa_id}")
    if st.button("💾 Salvar Informações", key=f"save_info_{empresa_id}", use_container_width=True):
        salvar_informacoes(empresa_id, info_texto)
        st.success("Informações salvas!")

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✏️ Editar Cadastro", use_container_width=True, key=f"edit_from_dialog_{empresa_id}"):
            st.session_state.edit_empresa_id = empresa_id
            st.rerun()
    with c2:
        if st.button("✕ Fechar", use_container_width=True, key=f"fechar_{empresa_id}"):
            st.rerun()


# ══════════════════════════════════════════════════════════════
# EMPRESAS — cards com progresso
# ══════════════════════════════════════════════════════════════
if menu == "📋 Empresas":
    st.markdown(f"### Empresas — {_MESES[mes_ref]}/{ano_ref}")

    col_s, col_f = st.columns([3, 1])
    with col_s:
        search = st.text_input("🔍 Buscar empresa", placeholder="Nome, código, CNPJ...")
    with col_f:
        filtro = st.selectbox("Filtrar", ["Todas", "Completas", "Em Progresso", "Não Iniciadas"])

    if search:
        empresas_prog = buscar_empresas_com_progresso(search, mes_str, ano_ref)
    else:
        empresas_prog = obter_empresas_com_progresso(mes_str, ano_ref)

    if filtro != "Todas":
        if filtro == "Completas":
            empresas_prog = [e for e in empresas_prog if pct_emp(e) == 100]
        elif filtro == "Em Progresso":
            empresas_prog = [e for e in empresas_prog if 0 < pct_emp(e) < 100]
        elif filtro == "Não Iniciadas":
            empresas_prog = [e for e in empresas_prog if pct_emp(e) == 0]

    faltando_map = obter_etapas_pendentes_bulk(mes_str, ano_ref)
    t_stats = obter_estatisticas_mes(mes_str, ano_ref)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total de Empresas", t_stats[0])
    m2.metric("✅ Completas",       t_stats[1])
    m3.metric("🔄 Em Progresso",   t_stats[2])
    m4.metric("⭕ Não Iniciadas",  t_stats[3])
    st.markdown("---")

    if not empresas_prog:
        st.info("Nenhuma empresa encontrada.")
    else:
        POR_PAG    = 12
        total_emp  = len(empresas_prog)
        total_pags = max(1, -(-total_emp // POR_PAG))

        pag_key = f"pag_{filtro}_{search}_{mes_str}_{ano_ref}"
        if st.session_state.get("_pag_ctx") != pag_key:
            st.session_state["_pag_ctx"] = pag_key
            st.session_state["pagina"]   = 1

        pagina = st.session_state.get("pagina", 1)
        pagina = max(1, min(pagina, total_pags))

        inicio    = (pagina - 1) * POR_PAG
        fim       = inicio + POR_PAG
        pagina_emp = empresas_prog[inicio:fim]

        _cap_col, _lote_col = st.columns([4, 1])
        with _cap_col:
            st.caption(f"Mostrando {inicio+1}–{min(fim, total_emp)} de **{total_emp}** empresa(s)  |  Página {pagina}/{total_pags}")
        with _lote_col:
            if st.button("⚡ Marcar Etapa em Lote", use_container_width=True, key="btn_lote_open"):
                dialog_marcar_lote()

        for emp in pagina_emp:
            pct  = pct_emp(emp)
            cor  = progress_color(pct)
            av   = initials(emp[1])
            faltando = faltando_map.get(emp[0], [])[:2]

            cod_label = f"Cód. {emp[10]} · " if emp[10] else ""
            sub_label = f"{cod_label}{_STATUS.get(emp[7], emp[7])}"
            if emp[11]: sub_label += f" · {emp[11]}"

            with st.container():
                st.markdown(f"""
                <div class="card">
                  <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:12px;">
                    <div style="display:flex;align-items:center;gap:12px;">
                      <div class="avatar">{av}</div>
                      <div>
                        <div class="company-name">{emp[1]}</div>
                        <div class="company-sub">{sub_label}</div>
                      </div>
                    </div>
                  </div>
                  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;">
                    <span style="font-size:12px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:.5px;">Progresso</span>
                    <span class="prog-pct" style="color:{cor};">{pct}%</span>
                  </div>
                  <div class="prog-wrap">
                    <div class="prog-bar" style="width:{pct}%;background:{cor};"></div>
                  </div>
                  {"<div style='margin-top:12px;'><span style='font-size:11px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:.5px;'>Faltando:</span><br>" + "".join(f'<span class="tag">⏳ {n}</span>' for n in faltando) + "</div>" if faltando else ""}
                </div>
                """, unsafe_allow_html=True)

                bc1, bc2, bc3, _ = st.columns([2, 1, 1, 4])
                with bc1:
                    if st.button("📂 Abrir Detalhes", key=f"open_{emp[0]}", use_container_width=True):
                        dialog_empresa(emp[0])
                with bc2:
                    if st.button("✏️", key=f"edit_card_{emp[0]}", help="Editar"):
                        st.session_state.edit_empresa_id = emp[0]
                with bc3:
                    key_c = f"confirm_del_{emp[0]}"
                    if st.session_state.get(key_c):
                        if st.button("✅", key=f"del_ok_{emp[0]}", help="Confirmar"):
                            deletar_empresa(emp[0])
                            st.session_state.pop(key_c, None)
                            st.rerun()
                    else:
                        if st.button("🗑️", key=f"del_{emp[0]}", help="Excluir"):
                            st.session_state[key_c] = True
                            st.rerun()

                if st.session_state.get(f"confirm_del_{emp[0]}"):
                    st.warning(f"Confirmar exclusão de **{emp[1]}**? Clique em ✅ acima.")

        if total_pags > 1:
            st.markdown("---")
            cols_pag = st.columns([1, 6, 1])
            with cols_pag[0]:
                if st.button("◀ Anterior", disabled=(pagina == 1), use_container_width=True):
                    st.session_state["pagina"] = pagina - 1
                    st.rerun()
            with cols_pag[1]:
                MAX_BTNS = 7
                if total_pags <= MAX_BTNS:
                    paginas_mostrar = list(range(1, total_pags + 1))
                else:
                    start_p = max(1, pagina - 2)
                    end_p   = min(total_pags, pagina + 2)
                    paginas_mostrar = list(range(start_p, end_p + 1))
                    if start_p > 1:        paginas_mostrar = [1, "..."] + paginas_mostrar
                    if end_p < total_pags: paginas_mostrar = paginas_mostrar + ["...", total_pags]

                btn_cols = st.columns(len(paginas_mostrar))
                for i, pg in enumerate(paginas_mostrar):
                    with btn_cols[i]:
                        if pg == "...":
                            st.markdown("<div style='text-align:center;padding-top:6px;color:#94a3b8;'>…</div>",
                                        unsafe_allow_html=True)
                        else:
                            label = f"**{pg}**" if pg == pagina else str(pg)
                            tipo  = "primary" if pg == pagina else "secondary"
                            if st.button(label, key=f"pag_btn_{pg}", type=tipo, use_container_width=True):
                                st.session_state["pagina"] = pg
                                st.rerun()
            with cols_pag[2]:
                if st.button("Próxima ▶", disabled=(pagina == total_pags), use_container_width=True):
                    st.session_state["pagina"] = pagina + 1
                    st.rerun()

    # formulário de edição inline
    if "edit_empresa_id" in st.session_state:
        emp = obter_empresa_por_id(st.session_state.edit_empresa_id)
        if emp:
            st.markdown("---")
            st.markdown("### ✏️ Editar Empresa")
            with st.form("form_edicao"):
                c1, c2 = st.columns(2)
                with c1:
                    nome   = st.text_input("Nome *",             value=emp[1])
                    codigo = st.text_input("Código",             value=safe(emp[10]))
                    cnpj   = st.text_input("CNPJ",               value=safe(emp[5]))
                    regime = st.text_input("Regime Tributário",  value=safe(emp[11]))
                    email  = st.text_input("E-mail",             value=safe(emp[3]))
                with c2:
                    responsavel = st.text_input("Responsável",   value=safe(emp[2]))
                    telefone    = st.text_input("Telefone",      value=safe(emp[4]))
                    endereco    = st.text_input("Endereço",      value=safe(emp[6]))
                    status = st.selectbox("Status", ["ativa","inativa","suspensa"],
                                          index=["ativa","inativa","suspensa"].index(emp[7]),
                                          format_func=lambda x: STATUS_PT[x])
                observacoes = st.text_area("Observações", value=safe(emp[9]), height=80)

                sc1, sc2 = st.columns(2)
                with sc1:
                    salvar = st.form_submit_button("✅ Salvar", use_container_width=True)
                with sc2:
                    cancelar = st.form_submit_button("❌ Cancelar", use_container_width=True)

                if cancelar:
                    del st.session_state.edit_empresa_id; st.rerun()
                if salvar:
                    ok, msg = atualizar_empresa(
                        st.session_state.edit_empresa_id,
                        nome, responsavel, email, telefone,
                        cnpj, endereco, status, observacoes, codigo, regime
                    )
                    if ok:
                        del st.session_state.edit_empresa_id; st.rerun()
                    else:
                        st.error(msg)


# ══════════════════════════════════════════════════════════════
# CADASTRAR EMPRESA
# ══════════════════════════════════════════════════════════════
elif menu == "➕ Cadastrar Empresa":
    st.markdown("### Cadastrar Nova Empresa")
    st.caption("As etapas padrão (Notas de Entrada/Saída/Serviço e Conciliação) serão adicionadas automaticamente.")

    with st.form("form_cadastro", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            nome   = st.text_input("Nome da Empresa *", placeholder="Razão Social")
            codigo = st.text_input("Código",             placeholder="Ex: 031")
            cnpj   = st.text_input("CNPJ",               placeholder="00.000.000/0000-00")
            regime = st.text_input("Regime Tributário",  placeholder="Simples / Presumido / Real")
        with c2:
            responsavel = st.text_input("Responsável",  placeholder="Nome completo")
            telefone    = st.text_input("Telefone",     placeholder="(11) 99999-9999")
            email       = st.text_input("E-mail",       placeholder="empresa@example.com")
            status      = st.selectbox("Status", ["ativa","inativa","suspensa"],
                                       format_func=lambda x: STATUS_PT[x])
        observacoes = st.text_area("Observações", height=80)

        if st.form_submit_button("✅ Cadastrar Empresa", use_container_width=True):
            if not nome.strip():
                st.error("Nome da empresa é obrigatório!")
            else:
                ok, msg = adicionar_empresa(nome.strip(), responsavel, email, telefone,
                                            cnpj, "", status, observacoes, codigo, regime)
                st.success(msg) if ok else st.error(msg)
                if ok: st.rerun()


# ══════════════════════════════════════════════════════════════
# IMPORTAR PLANILHA
# ══════════════════════════════════════════════════════════════
elif menu == "📥 Importar Planilha":
    st.markdown("### Importar Empresas via Planilha")

    arquivo = st.file_uploader("Selecione o arquivo (.xlsx, .xls, .csv)",
                               type=["xlsx","xls","csv"])

    if arquivo:
        try:
            is_csv = arquivo.name.endswith(".csv")

            aba_sel = None
            if not is_csv:
                import openpyxl
                wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
                abas = wb.sheetnames
                wb.close()
                arquivo.seek(0)
                aba_sel = st.selectbox("📄 Aba da planilha", abas) if len(abas) > 1 else abas[0]
                if len(abas) == 1:
                    st.caption(f"Aba: **{aba_sel}**")

            arquivo.seek(0)
            if is_csv:
                df = pd.read_csv(arquivo, dtype=str)
            else:
                df = pd.read_excel(arquivo, sheet_name=aba_sel, dtype=str)

            df.columns = df.columns.str.strip()
            df         = df.dropna(how="all").reset_index(drop=True)
            df         = df.replace(r'^\s*nan\s*$', '', regex=True).fillna('')

            st.info(f"📊 **{len(df)}** linhas encontradas no arquivo")

            def detectar(cands, cols):
                for cand in cands:
                    for col in cols:
                        if cand.lower() in col.lower(): return col
                return "— não usar —"

            opcoes     = ["— não usar —"] + list(df.columns)
            auto_nome  = detectar(["razão social","razao social","razao","nome","empresa"], df.columns)
            auto_cod   = detectar(["código","codigo","cod"], df.columns)
            auto_cnpj  = detectar(["cnpj"], df.columns)
            auto_reg   = detectar(["regime","tribut"], df.columns)
            auto_resp  = detectar(["responsavel","responsável"], df.columns)
            auto_tel   = detectar(["telefone","fone","celular"], df.columns)
            auto_email = detectar(["email","e-mail"], df.columns)

            st.markdown("#### Mapeamento de Colunas")
            c1, c2 = st.columns(2)
            with c1:
                col_nome  = st.selectbox("Coluna → Nome *",            opcoes, index=opcoes.index(auto_nome)  if auto_nome  in opcoes else 0)
                col_cod   = st.selectbox("Coluna → Código",            opcoes, index=opcoes.index(auto_cod)   if auto_cod   in opcoes else 0)
                col_cnpj  = st.selectbox("Coluna → CNPJ",              opcoes, index=opcoes.index(auto_cnpj)  if auto_cnpj  in opcoes else 0)
                col_reg   = st.selectbox("Coluna → Regime Tributário", opcoes, index=opcoes.index(auto_reg)   if auto_reg   in opcoes else 0)
            with c2:
                col_resp  = st.selectbox("Coluna → Responsável", opcoes, index=opcoes.index(auto_resp)  if auto_resp  in opcoes else 0)
                col_tel   = st.selectbox("Coluna → Telefone",    opcoes, index=opcoes.index(auto_tel)   if auto_tel   in opcoes else 0)
                col_email = st.selectbox("Coluna → E-mail",      opcoes, index=opcoes.index(auto_email) if auto_email in opcoes else 0)
                status_p  = st.selectbox("Status padrão", ["ativa","inativa","suspensa"],
                                         format_func=lambda x: STATUS_PT[x])

            if col_nome == "— não usar —":
                st.warning("Selecione a coluna **Nome da Empresa**.")
            else:
                nomes_serie = df[col_nome].str.strip()
                mask        = nomes_serie.ne("") & nomes_serie.str.lower().ne("nan") & nomes_serie.notna()
                df_val      = df[mask].copy()

                st.success(f"✅ **{len(df_val)}** empresa(s) prontas para importar")

                df_show = pd.DataFrame({"Nome": df_val[col_nome].values})
                if col_cod  != "— não usar —": df_show["Código"] = df_val[col_cod].str.strip().values
                if col_cnpj != "— não usar —": df_show["CNPJ"]   = df_val[col_cnpj].str.strip().values
                if col_reg  != "— não usar —": df_show["Regime"] = df_val[col_reg].str.strip().values
                st.dataframe(df_show, use_container_width=True, hide_index=True, height=280)

                if st.button("⚡ Importar TODAS as Empresas", type="primary", use_container_width=True):
                    def gcol(row, col):
                        return row[col].strip() if col != "— não usar —" and col in row.index else ""

                    registros = [
                        {
                            "nome":        row[col_nome].strip(),
                            "codigo":      gcol(row, col_cod),
                            "cnpj":        gcol(row, col_cnpj),
                            "regime":      gcol(row, col_reg),
                            "responsavel": gcol(row, col_resp),
                            "telefone":    gcol(row, col_tel),
                            "email":       gcol(row, col_email),
                            "observacoes": f"Importado em {datetime.today().strftime('%d/%m/%Y')}",
                        }
                        for _, row in df_val.iterrows()
                    ]

                    with st.spinner("Importando... aguarde."):
                        imp, dup, err = importar_empresas_bulk(registros, status_p)

                    if imp: st.success(f"✅ {imp} empresa(s) importada(s) com sucesso!")
                    if dup: st.warning(f"⚠️ {dup} ignorada(s) — já existiam no cadastro.")
                    if err: st.error(f"❌ {err} erro(s).")
                    if imp: st.rerun()

        except Exception as e:
            st.error(f"Erro ao ler o arquivo: {e}")
            st.exception(e)
    else:
        st.markdown("""
        **Formato aceito:**

        | Código | Razão Social | CNPJ | Regime Tributário |
        |--------|-------------|------|-------------------|
        | 031 | Empresa A Ltda | 00.000.000/0001-00 | Simples |

        ✅ As 4 etapas padrão são criadas automaticamente para cada empresa importada.
        """)


# ══════════════════════════════════════════════════════════════
# RELATÓRIOS
# ══════════════════════════════════════════════════════════════
elif menu == "📈 Relatórios":
    mes_nome = _MESES[mes_ref]
    st.markdown("### 📈 Relatórios")

    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Visão do Mês",
        "📋 Histórico por Empresa",
        "📥 Exportar",
        "🔍 Buscar por Empresa",
    ])

    # ── TAB 1: Visão do Mês ───────────────────────────────────
    with tab1:
        st.markdown(f"#### Resumo — {mes_nome}/{ano_ref}")

        # batch queries
        empresas_prog = obter_empresas_com_progresso(mes_str, ano_ref)
        t_stats       = obter_estatisticas_mes(mes_str, ano_ref)
        total_emp_r, completas_r, em_prog_r, nao_ini_r = t_stats

        if total_emp_r == 0:
            st.info("Nenhuma empresa cadastrada.")
        else:
            # métricas
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("🏢 Total",          total_emp_r)
            c2.metric("✅ Completas",       completas_r)
            c3.metric("🔄 Em Progresso",   em_prog_r)
            c4.metric("⭕ Não Iniciadas",  nao_ini_r)
            st.markdown("---")

            # charts
            col_pie, col_bar = st.columns([1, 2])
            paleta = PALETAS.get(cfg["daltonico"], PALETAS["normal"])

            with col_pie:
                fig_pie = go.Figure(go.Pie(
                    labels=["Completas", "Em Progresso", "Não Iniciadas"],
                    values=[completas_r, em_prog_r, nao_ini_r],
                    hole=.55,
                    marker_colors=[paleta["completo"], paleta["parcial"], paleta["vazio"]],
                    textinfo="label+percent",
                    hovertemplate="%{label}: %{value}<extra></extra>",
                ))
                fig_pie.update_layout(
                    showlegend=False,
                    margin=dict(l=10, r=10, t=10, b=10),
                    height=240,
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    font_color=_text_main,
                )
                st.plotly_chart(fig_pie, use_container_width=True, config={"displayModeBar": False})

            with col_bar:
                # top 10 by progress
                dados_chart = sorted(
                    [(e[1][:25], round(e[14]/e[13]*100) if e[13] > 0 else 0)
                     for e in empresas_prog],
                    key=lambda x: x[1], reverse=True
                )[:12]
                nomes_c = [d[0] for d in dados_chart]
                pcts_c  = [d[1] for d in dados_chart]
                cores_c = [paleta["completo"] if p == 100 else paleta["parcial"] if p >= 50 else paleta["vazio"]
                           for p in pcts_c]

                fig_bar = go.Figure(go.Bar(
                    x=pcts_c, y=nomes_c, orientation="h",
                    marker_color=cores_c,
                    text=[f"{p}%" for p in pcts_c],
                    textposition="outside",
                    hovertemplate="%{y}: %{x}%<extra></extra>",
                ))
                fig_bar.update_layout(
                    title="Top Empresas por Progresso",
                    xaxis=dict(range=[0, 115], ticksuffix="%"),
                    margin=dict(l=10, r=30, t=40, b=10),
                    height=300,
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    font_color=_text_main,
                    yaxis=dict(autorange="reversed"),
                )
                st.plotly_chart(fig_bar, use_container_width=True, config={"displayModeBar": False})

            st.markdown("---")
            st.markdown("#### Todas as Empresas")

            # tabela estilizada
            etapas_bulk = obter_todas_etapas_status_bulk(mes_str, ano_ref)

            rows_html = ""
            for emp in empresas_prog:
                pct  = round(emp[14]/emp[13]*100) if emp[13] > 0 else 0
                cor  = progress_color(pct)
                done = int(emp[14]) if emp[14] else 0
                tot  = int(emp[13]) if emp[13] else 0

                et_status = etapas_bulk.get(emp[0], {})
                tags = "".join(
                    f'<span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{"#10b981" if v else "#e2e8f0"};margin-right:3px;" title="{k}"></span>'
                    for k, v in et_status.items()
                )

                rows_html += f"""
                <tr>
                  <td style="font-weight:600;">{safe(emp[10],'-')}</td>
                  <td style="font-weight:600;">{emp[1]}</td>
                  <td style="color:{_text_sub}">{safe(emp[5],'-')}</td>
                  <td style="color:{_text_sub}">{safe(emp[11],'-')}</td>
                  <td>
                    <div style="display:flex;align-items:center;gap:8px;">
                      <div class="mini-prog-wrap">
                        <div class="mini-prog-bar" style="width:{pct}%;background:{cor};"></div>
                      </div>
                      <span style="font-weight:700;color:{cor};min-width:36px;">{pct}%</span>
                    </div>
                  </td>
                  <td style="color:{_text_sub}">{done}/{tot}</td>
                  <td>{tags}</td>
                </tr>"""

            st.markdown(f"""
            <div style="overflow-x:auto;max-height:480px;overflow-y:auto;border-radius:10px;border:1px solid {_border};">
            <table class="rel-table">
              <thead>
                <tr>
                  <th>Código</th><th>Empresa</th><th>CNPJ</th>
                  <th>Regime</th><th>Progresso</th><th>Etapas</th><th>Detalhes</th>
                </tr>
              </thead>
              <tbody>{rows_html}</tbody>
            </table>
            </div>
            """, unsafe_allow_html=True)

    # ── TAB 2: Histórico por Empresa ──────────────────────────
    with tab2:
        st.markdown("#### Histórico por Empresa")
        empresas_lista = obter_todas_empresas()
        if not empresas_lista:
            st.info("Nenhuma empresa cadastrada.")
        else:
            emp_sel = st.selectbox(
                "Selecione a empresa",
                empresas_lista,
                format_func=lambda x: f"{safe(x[10]+' · ') if x[10] else ''}{x[1]}",
                key="rel_emp_hist"
            )
            if emp_sel:
                c1, c2, c3 = st.columns(3)
                c1.write(f"**Código:** {safe(emp_sel[10],'N/A')}")
                c2.write(f"**CNPJ:** {safe(emp_sel[5],'N/A')}")
                c3.write(f"**Regime:** {safe(emp_sel[11],'N/A')}")

                st.markdown("---")

                # progresso anual em batch (1 query)
                prog_anual = obter_progresso_anual_bulk(emp_sel[0], ano_ref)

                meses_nomes = [_MESES[m] for m in range(1, 13)]
                pcts_anuais = [prog_anual.get(f"{m:02d}", (0, 0, 0))[0] for m in range(1, 13)]
                paleta = PALETAS.get(cfg["daltonico"], PALETAS["normal"])

                # gráfico de linha
                fig_line = go.Figure()
                fig_line.add_trace(go.Scatter(
                    x=meses_nomes, y=pcts_anuais,
                    mode="lines+markers+text",
                    line=dict(color="#3b82f6", width=3),
                    marker=dict(size=8, color=[
                        paleta["completo"] if p == 100 else paleta["parcial"] if p >= 50 else paleta["vazio"]
                        for p in pcts_anuais
                    ]),
                    text=[f"{p}%" for p in pcts_anuais],
                    textposition="top center",
                    hovertemplate="%{x}: %{y}%<extra></extra>",
                ))
                fig_line.add_hline(y=100, line_dash="dot", line_color=paleta["completo"],
                                   annotation_text="100%", annotation_position="right")
                fig_line.update_layout(
                    title=f"Progresso Mensal — {ano_ref}",
                    yaxis=dict(range=[0, 115], ticksuffix="%", gridcolor=_border),
                    xaxis=dict(gridcolor=_border),
                    height=340,
                    margin=dict(l=10, r=40, t=50, b=10),
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    font_color=_text_main,
                )
                st.plotly_chart(fig_line, use_container_width=True, config={"displayModeBar": False})

                # tabela mensal
                etapas_emp = obter_etapas_empresa(emp_sel[0])
                hist_rows = []
                for m in range(1, 13):
                    ms       = f"{m:02d}"
                    prog     = prog_anual.get(ms, (0, 0, 0))
                    stat_map = obter_status_etapas(emp_sel[0], ms, ano_ref)
                    row      = {"Mês": _MESES[m], "Progresso": f"{prog[0]}%", "Concluídas": f"{prog[1]}/{prog[2]}"}
                    for et in etapas_emp:
                        row[et[2]] = "✅" if stat_map.get(et[0]) else "❌"
                    hist_rows.append(row)

                df_hist = pd.DataFrame(hist_rows)
                st.dataframe(
                    df_hist,
                    use_container_width=True,
                    hide_index=True,
                    height=460,
                )

                # ── Exportar histórico desta empresa ──────────
                st.markdown("---")
                st.markdown("##### ⬇️ Exportar Histórico desta Empresa")
                nome_emp_safe = emp_sel[1].replace("/", "_").replace("\\", "_")[:40]

                # ── prepara arquivos FORA dos contextos de coluna ──
                # Excel
                _buf_h = io.BytesIO()
                with pd.ExcelWriter(_buf_h, engine="openpyxl") as _w:
                    df_hist.to_excel(_w, sheet_name="Historico", index=False)
                _excel_bytes = _buf_h.getvalue()

                # CSV
                _csv_h = df_hist.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

                # PDF (gerado fora de qualquer `with col:`)
                _pdf_bytes = None
                if HAS_FPDF:
                    def _build_pdf_hist(df, empresa, cod, cnpj, regime, ano):
                        def _ps(t, n=35):
                            return str(t)[:n].encode("latin-1", errors="replace").decode("latin-1")
                        _col_names  = list(df.columns)
                        _n_extra    = max(len(_col_names) - 3, 0)

                        p = FPDF(orientation="L", unit="mm", format="A4")
                        p.set_auto_page_break(auto=True, margin=15)
                        p.add_page()

                        # larguras das colunas — calculadas com p.epw para não vazar
                        _W_MES  = 26   # Mês
                        _W_PCT  = 24   # Progresso
                        _W_CONC = 26   # Concluídas
                        _fixed_w = _W_MES + _W_PCT + _W_CONC   # 76 mm
                        if _n_extra > 0:
                            _extra_w = max(20, min(60, int((p.epw - _fixed_w) / _n_extra)))
                        else:
                            _extra_w = 0
                        _widths = [_W_MES, _W_PCT, _W_CONC] + [_extra_w] * _n_extra

                        # cabeçalho azul
                        p.set_fill_color(59, 130, 246)
                        p.rect(0, 0, 297, 28, "F")
                        p.set_font("Helvetica", "B", 16)
                        p.set_text_color(255, 255, 255)
                        p.set_xy(10, 5)
                        p.cell(0, 9, _ps(f"Historico Anual - {empresa}", 55), new_x="LMARGIN", new_y="NEXT")
                        p.set_font("Helvetica", "", 9)
                        p.set_xy(10, 17)
                        info_line = (f"Ano: {ano}   Codigo: {_ps(cod)}   "
                                     f"CNPJ: {_ps(cnpj)}   Regime: {_ps(regime)}")
                        p.cell(200, 6, info_line)
                        p.set_xy(10, 17)
                        p.cell(0, 6, f"Gerado: {datetime.today().strftime('%d/%m/%Y %H:%M')}", align="R")

                        p.set_text_color(0, 0, 0)
                        p.set_y(34)

                        # cabeçalho da tabela
                        p.set_font("Helvetica", "B", 9)
                        p.set_fill_color(59, 130, 246)
                        p.set_text_color(255, 255, 255)
                        for _cn, _wn in zip(_col_names, _widths):
                            # caracteres que cabem ≈ largura_mm / 2  (fonte 9pt ≈ 2mm/char)
                            _max_c = max(6, int(_wn / 2))
                            p.cell(_wn, 8, _ps(_cn, _max_c), border=1, fill=True)
                        p.ln()

                        # linhas de dados
                        p.set_font("Helvetica", "", 8)
                        p.set_text_color(0, 0, 0)
                        for _i, _row in df.iterrows():
                            if _i % 2 == 0:
                                p.set_fill_color(248, 250, 252)
                            else:
                                p.set_fill_color(255, 255, 255)
                            for _cn, _wn in zip(_col_names, _widths):
                                _val = str(_row[_cn]).replace("✅", "Sim").replace("❌", "Nao")
                                _max_c = max(6, int(_wn / 2))
                                p.cell(_wn, 7, _ps(_val, _max_c), border=1, fill=True)
                            p.ln()

                        return bytes(p.output())

                    _pdf_bytes = _build_pdf_hist(
                        df_hist,
                        emp_sel[1],
                        safe(emp_sel[10], "N/A"),
                        safe(emp_sel[5],  "N/A"),
                        safe(emp_sel[11], "N/A"),
                        ano_ref,
                    )

                # ── botões de download dentro das colunas ──
                hc1, hc2, hc3 = st.columns(3)
                with hc1:
                    st.download_button(
                        "📊 Baixar Excel",
                        _excel_bytes,
                        file_name=f"historico_{nome_emp_safe}_{ano_ref}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="dl_hist_xlsx",
                    )
                with hc2:
                    st.download_button(
                        "📄 Baixar CSV",
                        _csv_h,
                        file_name=f"historico_{nome_emp_safe}_{ano_ref}.csv",
                        mime="text/csv",
                        use_container_width=True,
                        key="dl_hist_csv",
                    )
                with hc3:
                    if not HAS_FPDF:
                        st.warning("pip install fpdf2")
                    else:
                        st.download_button(
                            "📑 Baixar PDF",
                            _pdf_bytes,
                            file_name=f"historico_{nome_emp_safe}_{ano_ref}.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key="dl_hist_pdf",
                        )

    # ── TAB 3: Exportar ───────────────────────────────────────
    with tab3:
        st.markdown(f"#### Exportar Dados — {mes_nome}/{ano_ref}")

        empresas_prog_exp = obter_empresas_com_progresso(mes_str, ano_ref)
        etapas_bulk_exp   = obter_todas_etapas_status_bulk(mes_str, ano_ref)

        # monta dataframe de relatório
        rows_exp = []
        for emp in empresas_prog_exp:
            pct  = round(emp[14]/emp[13]*100) if emp[13] > 0 else 0
            done = int(emp[14]) if emp[14] else 0
            tot  = int(emp[13]) if emp[13] else 0
            row  = {
                "Código":      safe(emp[10], "-"),
                "Empresa":     emp[1],
                "CNPJ":        safe(emp[5], "-"),
                "Regime":      safe(emp[11], "-"),
                "Status":      _STATUS.get(emp[7], emp[7]),
                "Progresso":   f"{pct}%",
                "Concluídas":  f"{done}/{tot}",
            }
            for nome_et, feito in etapas_bulk_exp.get(emp[0], {}).items():
                row[nome_et] = "Sim" if feito else "Não"
            rows_exp.append(row)
        df_exp = pd.DataFrame(rows_exp)

        # preview
        if not df_exp.empty:
            st.dataframe(df_exp, use_container_width=True, hide_index=True, height=300)

        st.markdown("---")
        col_x, col_c, col_p = st.columns(3)

        # ── Excel ──
        with col_x:
            st.markdown("##### 📊 Excel")
            buf_xl = io.BytesIO()
            with pd.ExcelWriter(buf_xl, engine="openpyxl") as w:
                df_exp.to_excel(w, sheet_name=f"{mes_nome} {ano_ref}", index=False)
                # Empresas completo
                obter_dataframe_empresas().to_excel(w, sheet_name="Cadastro Completo", index=False)
            st.download_button(
                "⬇️ Baixar Excel",
                buf_xl.getvalue(),
                file_name=f"relatorio_{mes_nome}_{ano_ref}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # ── CSV ──
        with col_c:
            st.markdown("##### 📄 CSV")
            csv_bytes = df_exp.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button(
                "⬇️ Baixar CSV",
                csv_bytes,
                file_name=f"relatorio_{mes_nome}_{ano_ref}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        # ── PDF ──
        with col_p:
            st.markdown("##### 📑 PDF")
            if not HAS_FPDF:
                st.warning("Instale **fpdf2** para exportar PDF:")
                st.code("pip install fpdf2", language="bash")
            elif df_exp.empty:
                st.info("Sem dados para exportar.")
            else:
                t_s = obter_estatisticas_mes(mes_str, ano_ref)

                def gerar_pdf():
                    def ps(text, n=40):
                        """Limita texto e codifica para Latin-1."""
                        t = str(text)[:n] if text else "-"
                        return t.encode("latin-1", errors="replace").decode("latin-1")

                    pdf = FPDF(orientation="L", unit="mm", format="A4")
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf.add_page()

                    # ── Cabeçalho ──
                    pdf.set_fill_color(59, 130, 246)
                    pdf.rect(0, 0, 297, 28, "F")
                    pdf.set_font("Helvetica", "B", 20)
                    pdf.set_text_color(255, 255, 255)
                    pdf.set_xy(10, 6)
                    pdf.cell(0, 10, ps(f"Relatorio de Empresas - {mes_nome}/{ano_ref}", 60), ln=True)
                    pdf.set_font("Helvetica", "", 10)
                    pdf.set_xy(10, 18)
                    pdf.cell(0, 6, f"Gerado em {datetime.today().strftime('%d/%m/%Y %H:%M')}")

                    pdf.set_text_color(0, 0, 0)
                    pdf.set_y(34)

                    # ── Resumo ──
                    pdf.set_font("Helvetica", "B", 11)
                    pdf.cell(0, 8, "RESUMO DO MES", ln=True)
                    pdf.set_font("Helvetica", "", 10)
                    pdf.set_fill_color(241, 245, 249)
                    pdf.cell(70, 8, f"Total de Empresas: {t_s[0]}", fill=True, border=1)
                    pdf.cell(70, 8, f"Completas: {t_s[1]}",        fill=True, border=1)
                    pdf.cell(70, 8, f"Em Progresso: {t_s[2]}",     fill=True, border=1)
                    pdf.cell(67, 8, f"Nao Iniciadas: {t_s[3]}",    fill=True, border=1)
                    pdf.ln(12)

                    # ── Tabela ──
                    pdf.set_font("Helvetica", "B", 9)
                    pdf.set_fill_color(59, 130, 246)
                    pdf.set_text_color(255, 255, 255)
                    cols_pdf  = ["Codigo", "Empresa", "CNPJ", "Regime", "Status", "Progresso", "Etapas"]
                    widths_pdf = [18, 70, 34, 30, 20, 22, 20]
                    for col_h, w_h in zip(cols_pdf, widths_pdf):
                        pdf.cell(w_h, 8, col_h, border=1, fill=True)
                    pdf.ln()

                    pdf.set_font("Helvetica", "", 8)
                    pdf.set_text_color(0, 0, 0)
                    for i, row in enumerate(rows_exp):
                        fill = (i % 2 == 0)
                        if fill:
                            pdf.set_fill_color(248, 250, 252)
                        else:
                            pdf.set_fill_color(255, 255, 255)
                        pct_v = row.get("Progresso", "0%")
                        conc  = row.get("Concluídas", "-")
                        vals  = [
                            ps(row.get("Código",  "-"), 8),
                            ps(row.get("Empresa", "-"), 35),
                            ps(row.get("CNPJ",    "-"), 18),
                            ps(row.get("Regime",  "-"), 15),
                            ps(row.get("Status",  "-"), 10),
                            ps(pct_v, 6),
                            ps(conc,  6),
                        ]
                        for val, w_r in zip(vals, widths_pdf):
                            pdf.cell(w_r, 7, val, border=1, fill=fill)
                        pdf.ln()

                    return bytes(pdf.output())

                pdf_bytes = gerar_pdf()
                st.download_button(
                    "⬇️ Baixar PDF",
                    pdf_bytes,
                    file_name=f"relatorio_{mes_nome}_{ano_ref}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )

    # ── TAB 4: Buscar por Empresa ─────────────────────────────
    with tab4:
        st.markdown("#### 🔍 Buscar por Empresa")
        st.caption(f"Período: **{mes_nome}/{ano_ref}** — use os seletores na barra lateral para mudar")

        busca_rel = st.text_input(
            "Digite nome, código ou CNPJ",
            placeholder="Ex: 031 | Empresa X | 00.000.000/0001-00",
            key="busca_relatorio_tab4"
        )

        if busca_rel.strip():
            resultados = buscar_empresas_com_progresso(busca_rel.strip(), mes_str, ano_ref)

            if not resultados:
                st.info("Nenhuma empresa encontrada para esse termo.")
            else:
                st.success(f"**{len(resultados)}** empresa(s) encontrada(s)")
                paleta_b = PALETAS.get(cfg["daltonico"], PALETAS["normal"])

                for emp_r in resultados:
                    pct_r  = round(emp_r[14]/emp_r[13]*100) if emp_r[13] > 0 else 0
                    cor_r  = progress_color(pct_r)
                    done_r = int(emp_r[14]) if emp_r[14] else 0
                    tot_r  = int(emp_r[13]) if emp_r[13] else 0

                    cod_r  = f"Cód. {emp_r[10]} · " if emp_r[10] else ""
                    sub_r  = f"{cod_r}{_STATUS.get(emp_r[7], emp_r[7])}"
                    if emp_r[11]: sub_r += f" · {emp_r[11]}"

                    with st.expander(f"**{emp_r[1]}** — {pct_r}%  ({done_r}/{tot_r} etapas)", expanded=True):
                        ca, cb = st.columns([3, 1])
                        with ca:
                            st.markdown(f"""
                            <div style="color:{_text_sub};margin-bottom:8px;">{sub_r}</div>
                            <div style="display:flex;align-items:center;gap:10px;">
                              <div style="flex:1;background:{_border};border-radius:4px;height:10px;">
                                <div style="width:{pct_r}%;background:{cor_r};height:10px;border-radius:4px;"></div>
                              </div>
                              <span style="font-weight:700;color:{cor_r};font-size:18px;">{pct_r}%</span>
                            </div>
                            """, unsafe_allow_html=True)

                        with cb:
                            if emp_r[5]:
                                st.caption(f"CNPJ: {emp_r[5]}")
                            if emp_r[3]:
                                st.caption(f"✉️ {emp_r[3]}")

                        # etapas do mês
                        et_status_r = obter_status_etapas(emp_r[0], mes_str, ano_ref)
                        etapas_r    = obter_etapas_empresa(emp_r[0])

                        if etapas_r:
                            st.markdown("---")
                            e_cols = st.columns(min(len(etapas_r), 4))
                            for idx_e, etapa_r in enumerate(etapas_r):
                                with e_cols[idx_e % len(e_cols)]:
                                    feito_r = et_status_r.get(etapa_r[0], False)
                                    icon    = "✅" if feito_r else "❌"
                                    st.markdown(
                                        f'<div style="text-align:center;padding:8px 4px;background:{_bg_card};border:1px solid {_border};border-radius:8px;margin-bottom:6px;">'
                                        f'<div style="font-size:20px;">{icon}</div>'
                                        f'<div style="font-size:11px;color:{_text_sub};margin-top:4px;">{etapa_r[2]}</div>'
                                        f'</div>',
                                        unsafe_allow_html=True
                                    )

                        if st.button("📂 Abrir Detalhes Completos", key=f"open_r_{emp_r[0]}"):
                            dialog_empresa(emp_r[0])
        else:
            st.markdown(f"""
            <div style="text-align:center;padding:3rem 2rem;color:{_text_sub};">
              <div style="font-size:48px;margin-bottom:1rem;">🔍</div>
              <div style="font-size:18px;font-weight:600;margin-bottom:.5rem;">Pesquise uma empresa</div>
              <div style="font-size:14px;">Digite o nome, código ou CNPJ para ver o detalhamento de etapas e progresso</div>
            </div>
            """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# CONFIGURAÇÕES
# ══════════════════════════════════════════════════════════════
elif menu == "⚙️ Configurações":
    st.markdown("### ⚙️ Configurações")
    st.caption("As configurações são aplicadas imediatamente e salvas na sessão atual.")

    # ── Tema ──────────────────────────────────────────────────
    st.markdown("#### 🎨 Aparência")
    col_t1, col_t2 = st.columns([2, 1])
    with col_t1:
        tema_opcoes = {"claro": "☀️ Modo Claro", "escuro": "🌙 Modo Escuro"}
        tema_novo = st.radio(
            "Tema",
            list(tema_opcoes.keys()),
            index=list(tema_opcoes.keys()).index(cfg["tema"]),
            format_func=lambda x: tema_opcoes[x],
            horizontal=True,
            key="cfg_tema"
        )
        if tema_novo != cfg["tema"]:
            st.session_state.cfg["tema"] = tema_novo
            st.rerun()

    # ── Tamanho da Fonte ──────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 🔤 Tamanho da Letra")
    fonte_opcoes = {"pequeno": "A  Pequeno (13px)", "medio": "A  Médio (15px)", "grande": "A  Grande (18px)"}
    fonte_nova = st.radio(
        "Tamanho",
        list(fonte_opcoes.keys()),
        index=list(fonte_opcoes.keys()).index(cfg["fonte"]),
        format_func=lambda x: fonte_opcoes[x],
        horizontal=True,
        key="cfg_fonte"
    )
    if fonte_nova != cfg["fonte"]:
        st.session_state.cfg["fonte"] = fonte_nova
        st.rerun()

    # ── Tema para Daltonismo ──────────────────────────────────
    st.markdown("---")
    st.markdown("#### 👁️ Acessibilidade — Daltonismo")
    st.caption("Ajusta as cores de progresso para melhor visibilidade conforme o tipo de daltonismo.")

    daltonico_opcoes = {
        "normal":       "🔵 Padrão",
        "deuteranopia": "🟡 Deuteranopia (vermelho-verde)",
        "protanopia":   "🟠 Protanopia (ausência de vermelho)",
        "tritanopia":   "🟣 Tritanopia (ausência de azul)",
    }
    dal_novo = st.radio(
        "Tema de cores",
        list(daltonico_opcoes.keys()),
        index=list(daltonico_opcoes.keys()).index(cfg["daltonico"]),
        format_func=lambda x: daltonico_opcoes[x],
        key="cfg_daltonico"
    )
    if dal_novo != cfg["daltonico"]:
        st.session_state.cfg["daltonico"] = dal_novo
        st.rerun()

    # preview das cores
    paleta_prev = PALETAS.get(dal_novo, PALETAS["normal"])
    st.markdown(
        f'<div style="display:flex;gap:16px;margin-top:12px;">'
        f'<div style="display:flex;align-items:center;gap:6px;"><div style="width:16px;height:16px;border-radius:50%;background:{paleta_prev["completo"]};"></div><span style="font-size:13px;">Completo (100%)</span></div>'
        f'<div style="display:flex;align-items:center;gap:6px;"><div style="width:16px;height:16px;border-radius:50%;background:{paleta_prev["parcial"]};"></div><span style="font-size:13px;">Em progresso (50–99%)</span></div>'
        f'<div style="display:flex;align-items:center;gap:6px;"><div style="width:16px;height:16px;border-radius:50%;background:{paleta_prev["vazio"]};"></div><span style="font-size:13px;">Não iniciado (0–49%)</span></div>'
        f'</div>',
        unsafe_allow_html=True
    )

    # ── Idioma ────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 🌐 Idioma / Language")
    idioma_opcoes = {"pt": "🇧🇷 Português", "en": "🇺🇸 English"}
    idioma_novo = st.radio(
        "Idioma",
        list(idioma_opcoes.keys()),
        index=list(idioma_opcoes.keys()).index(cfg["idioma"]),
        format_func=lambda x: idioma_opcoes[x],
        horizontal=True,
        key="cfg_idioma"
    )
    if idioma_novo != cfg["idioma"]:
        st.session_state.cfg["idioma"] = idioma_novo
        st.rerun()

    # ── Atalhos / Info ────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### ℹ️ Sobre o Sistema")
    st.markdown(f"""
    <div style="background:{_bg_card};border:1px solid {_border};border-radius:10px;padding:16px;">
      <p style="margin:4px 0;color:{_text_sub};font-size:14px;">🏢 <strong>Gestor de Empresas</strong></p>
      <p style="margin:4px 0;color:{_text_sub};font-size:13px;">Data atual: {today.strftime('%d/%m/%Y')}</p>
      <p style="margin:4px 0;color:{_text_sub};font-size:13px;">Tema ativo: {tema_opcoes[cfg['tema']]}</p>
      <p style="margin:4px 0;color:{_text_sub};font-size:13px;">Exportação PDF: {"✅ Disponível (fpdf2)" if HAS_FPDF else "❌ Instalar fpdf2 — pip install fpdf2"}</p>
    </div>
    """, unsafe_allow_html=True)

    if not HAS_FPDF:
        st.markdown("---")
        st.info("Para habilitar exportação em PDF, execute no terminal:")
        st.code("pip install fpdf2", language="bash")
